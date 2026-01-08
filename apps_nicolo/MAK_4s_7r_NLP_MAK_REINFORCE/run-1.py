# %%
# Set up the file path
import sys
import os
parent_dir = os.path.abspath(os.path.join(os.getcwd()))
sys.path.append(parent_dir)
task_name = 'MAK_3s_5r_NLP_MAK_REINFORCE_SIL'
print('Working directory set to:', parent_dir)

# %%
# Import general packages
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import torch
from matplotlib import pyplot as plt
from pytorch_lightning.loggers import CometLogger
import numpy as np
from itertools import product
from tqdm import tqdm

# Import Agent-Environment packages
from RL4CRN.environments.environment import Environment
from RL4CRN.environments.parallel_environments import ParallelEnvironments
from RL4CRN.environments.serial_environments import SerialEnvironments
from RL4CRN.agents.reinforce_agent import REINFORCEAgent
from RL4CRN.policies.add_reaction_by_ordered_index import AddReactionByOrderedIndex
from RL4CRN.policies.add_reaction_by_index import AddReactionByIndex

# Import Interface packages
from RL4CRN.env2agent_interface.explicit_observer import ExplicitObserver
from RL4CRN.env2agent_interface.explicit_tensorizer import ExplicitTensorizer
from RL4CRN.agent2env_interface.library_actuator import LibraryActuator
from RL4CRN.agent2env_interface.iocrn_stepper import IOCRNStepper

# Import CRN packages
from RL4CRN.iocrns.iocrn import IOCRN
from RL4CRN.iocrns.reactions import MassAction
from RL4CRN.utils.ic import IC
from RL4CRN.iocrns.reaction_library import construct_mass_action_library

# Import Reward packages
from RL4CRN.rewards.deterministic import dynamic_tracking_error

# Import Gemini Interface packages
from RL4CRN.NLPAgent.VertexMultiAgentDebate import VertexMultiAgentDebate
import time
os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "/local0/home/rossin/.keys/crn-evolution-be2b980ea837.json"

# %%
# Set the logger to use Comet
timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
api_key = "vhIR3uyqsKyU4L7SA8fLCfTSC"
logger = CometLogger(
    api_key=api_key,
    project=task_name,        
    workspace="redsnic", 
    name=f'{task_name}_{timestamp}',
)
logger = logger.experiment

# %%
# Construct the template CRN
scale = 1.0
r1 = MassAction(reactant_labels=[], product_labels=['Z_1'], input_channels=['u_1'], params=[scale], params_controllability=[True])
r2 = MassAction(reactant_labels=['X_1'], product_labels=[], input_channels=['u_2'], params=[1.], params_controllability=[True])
crn_template = IOCRN([r1, r2], output_labels=['X_1'])
crn_template.compile()
p = crn_template.num_inputs # Number of inputs of the IOCRNs
print("Template CRN:")
print(crn_template)

# Construct the library of possible reactions
species_labels = ['X_1', 'Z_1', 'Z_2']
library = construct_mass_action_library(species_labels=species_labels, order=2)
crn_template.set_library_context(library)
M = len(library.reactions) # Number of possible reactions
K = library.get_num_parameters() # Total number of parameters in all the reactions of the library
print("Library of possible reactions:")
print(library)
print("------------------------------------------------")

# %%
# Device
device = 'cuda' if torch.cuda.is_available() else 'cpu'
print(f'Using device: {device}')
print(f'Number of CPUs available: {os.cpu_count()}') 

# %%
# Flags and filenames
save_flag = True                                                # Save the agent checkpoint
load_flag = False                                               # Load the agent checkpoint 
train_flag = True                                               # Train the agent
save_sheet_flag = True                                          # Save the configuration to an Excel sheet

save_filename = timestamp + '.pth'                              # Filename for saving the agent checkpoint
load_filename = ''                                              # Filename for loading the agent checkpoint
file_name = f"{task_name}.xlsx"             # Filename for saving the Excel sheet

# %%
# %%
from RL4CRN.utils.visualizations import plot_truth_table

# Hyperparameters
max_added_reactions = 5                             # Maximum number of reactions
N_CPUs = os.cpu_count()                             # Number of CPUs          
N = 10*N_CPUs                                       # Number of samples (batch size)    
width = 1024                                        # Width of the neural networks  
depth = 5                                           # Depth of the neural networks 
deep_layer_size = 1024*10                           # Size of the deep layer encoding the CRNs
allow_input_influence = False                       # Allow input influence in the policy
learning_rate = 1e-4                                # Learning rate for the optimizer 
hall_of_fame_size = 100                              # Size of the hall of fame  
entropy_scheduler = {                               # Entropy scheduler parameters 
    'entropy_weight': 1e-3, 
    'topk_entropy_weight' : 1.0,
    'remainder_entropy_weight' : 1.0,
    'entropy_update_coefficient': 1, 
    'entropy_schedule': 1000, 
    'minimum_entropy_weight': 0.0
}
entropy_weights_per_head = {'structure': 2.0, 'continuous': 1.0, 'discrete': 0.0, 'input_influence': 0.0} 
structure_head_temperature = {"target_entropy_ratio_to_max": np.log(5)/np.log(M), "initial_temperature": 1.0, "rate": 0.0, "current_temperature": 1.0}
risk_scheduler = {                                  # Risk scheduler parameters
    'risk': 0.9, 
    'risk_update': 0.0, 
    'max_risk': 1.0, 
    'risk_schedule': 1000
}
epoch_num = 300                                     # Number of epochs for training
render_schedule = 10                                 # Render every # of epochs
render_mode = {                                     # Mode of the experiment
    'style': 'logger', 
    'task': 'transients', 
    'format': 'image',
    'topology': True,
    'bounds': [2.5]
}
# Ordering specific parameters
ordering_parameters = {
    'enforce_ordering': False,
    'constraint_weight' : float('inf')
}
# SIL settings
sil_settings = {
    'sil_loss_weight': 1.0,
    'sil_use_adaptive_baseline': False,
    'sil_baseline_annealing_rate': 0.95
}

render_n_best = 10                                                    # Number of best CRNs to plot responses for
render_disregard_percentage = 0.99                                    # Percentage of worst CRNs to disregard in the responses plotting

# Parameter distribution for the reactions added by the agent
continuous_distribution = {"type": 'lognormal_1D'}

# Time horizon for the simulation
t_f = 100                                           # Final time for the simulation
N_t = 1000                                          # Number of time steps
time_horizon = np.linspace(0, t_f, N_t, dtype=np.float32)

# Construct the IOCRN inputs
nums = [0.5, 1.0, 1.5]
u_list = [np.array(u) for u in product(nums, repeat=p)] # list of input combinations, each input is a numpy array of shape (p,)

# Construct the reference setpoints
r_list = [np.array([u[0]]) for u in u_list]

# Construct the IOCRN initial conditions
ic = IC(names=species_labels, values=[[0.0, 0.0, 0.0, 0.0]])

# Construct the weights for the performance metric
w = np.ones(N_t)
w[(len(w)//5)*4:] = w[(len(w)//5)*4:]*2
w[:(len(w)//5)] = w[:(len(w)//5)]*0.25
w = w[np.newaxis, :]

# Construct the compute reward routine
def compute_reward(state):
    x0_list = ic.get_ic(state)
    return dynamic_tracking_error(state, u_list, x0_list, time_horizon, r_list, w, norm=1, LARGE_NUMBER=1e4)

# %%
if save_sheet_flag:
    sheet_name = "Data"
    headers = [
        "Timestamp", "URL",
        "Epochs Completed", "Successful", "Saved", "Comments",
        "Learning Rate", "Epochs #",
        "(m, n, p, N)",
        "NN Depth", "NN Width", "Deep Layer Size", "CPUs #",
        "Entropy Scheduler",
        "Risk Scheduler",
        "Render Schedule", "HoF Size",
        "Simulation Time", "Time Steps #",
        "Initial Conditions #", "Input Scenarios#",
        "Continuous Distribution", "Entropy Weights per Head",
        "Structure Head Temperature",
        "Ordering Enforced",
        "SIL Settings"
    ]

    data_row = [
        timestamp, logger.url,
        None, None, None, None,
        learning_rate, epoch_num,
        str((max_added_reactions, len(species_labels), p, N)),
        depth, width, deep_layer_size, N_CPUs,
        str(entropy_scheduler),
        str(risk_scheduler),
        render_schedule, hall_of_fame_size,
        t_f, N_t, len(ic.values), len(u_list),
        str(continuous_distribution), str(entropy_weights_per_head),
        str(structure_head_temperature),
        f"Yes: {ordering_parameters['constraint_weight']}" if ordering_parameters['enforce_ordering'] else "No",
        str(sil_settings)
    ]

    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    # Write headers if sheet is empty
    if ws.max_row == 1 and ws.max_column == 1 and ws.cell(row=1, column=1).value is None:
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

    # Append experiment as next row
    next_row = ws.max_row + 1
    for col, value in enumerate(data_row, start=1):
        ws.cell(row=next_row, column=col, value=value)

    # Freeze header row and add filter
    ws.freeze_panes = "B1" 
    ws.auto_filter.ref = ws.dimensions

    # === Auto-fit column widths (except URL column) ===
    # URL column is column 2 (B), we leave its width unchanged.
    url_col_index = 2

    for col in range(1, ws.max_column + 1):
        if col == url_col_index:
            continue  # keep URL column width as-is

        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            value = cell.value
            if value is not None:
                # Convert to string to measure length
                length = len(str(value))
                if length > max_length:
                    max_length = length

        # Some padding so text isn't touching the cell border
        adjusted_width = max_length + 2 if max_length > 0 else 10
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(file_name)
    print(f"New experiment data saved in row {next_row} of '{file_name}'.")

# %%
# Construct parallel environments
crn_0 = crn_template.clone()
mult_env = ParallelEnvironments([Environment(crn_0, max_added_reactions, logger=logger, logger_schedule=1) for _ in range(N)], hall_of_fame_size=hall_of_fame_size, N_CPUs=N_CPUs, logger=logger)
# mult_env = SerialEnvironments([Environment(crn_0, max_added_reactions, logger=logger, logger_schedule=1) for _ in range(N)], hall_of_fame_size=hall_of_fame_size, logger=logger)

# %%
# Construct the policy
encoder_attributes = {"hidden_size": width, "num_layers": depth}
structure_head_attributes = {"hidden_size": width, "num_layers": depth}
rate_head_attributes = {"hidden_size": width, "num_layers": depth}
input_influence_head_attributes = {"hidden_size": width, "num_layers": depth}
masks = {"continuous": library.get_parameter_mask(mode="continuous"), "discrete": library.get_parameter_mask(mode="discrete"), "logit": library.get_logit_mask()}
policy = AddReactionByOrderedIndex(M, K, p, encoder_attributes, deep_layer_size, structure_head_attributes, rate_head_attributes, input_influence_head_attributes, target_set_size=crn_template.num_reactions+max_added_reactions, allow_input_influence=False, masks=masks, device=device, continuous_distribution=continuous_distribution, entropy_weights_per_head=entropy_weights_per_head,
                                    combinatorial_bias_enabled=ordering_parameters["enforce_ordering"], constraint_strength=ordering_parameters["constraint_weight"])

if ordering_parameters["enforce_ordering"]:
    policy = AddReactionByOrderedIndex(M, K, p, encoder_attributes, deep_layer_size, structure_head_attributes, rate_head_attributes, input_influence_head_attributes, target_set_size=crn_template.num_reactions+max_added_reactions, allow_input_influence=False, masks=masks, device=device, continuous_distribution=continuous_distribution, entropy_weights_per_head=entropy_weights_per_head,
                                        combinatorial_bias_enabled=ordering_parameters["enforce_ordering"], constraint_strength=ordering_parameters["constraint_weight"])
else:
    policy = AddReactionByIndex(M, K, p, encoder_attributes, deep_layer_size, structure_head_attributes, rate_head_attributes, input_influence_head_attributes, allow_input_influence=False, masks=masks, device=device, continuous_distribution=continuous_distribution, entropy_weights_per_head=entropy_weights_per_head)

# Construct the agent
agent = REINFORCEAgent(policy, allow_input_influence=False, logger=logger, learning_rate=learning_rate, entropy_scheduler=entropy_scheduler, risk_scheduler=risk_scheduler, sil_settings=sil_settings, device=device)
if load_flag:
    agent.policy.load_state_dict(torch.load(load_filename+'.pth', map_location=device))

# %%
# Construct the interfaces
observer = ExplicitObserver(reaction_library=library, allow_input_observation=allow_input_influence)
tensorizer = ExplicitTensorizer(device=device)
actuator = LibraryActuator(reaction_library=library)
stepper = IOCRNStepper()

# %%
# ==========================================
# 0. SETUP & UTILS
# ==========================================
# (Assuming imports for VertexMultiAgentDebate, Environment, etc. are already done)
# (Assuming 'agent', 'mult_env', 'library', 'stepper', 'actuator', 'observer', 'tensorizer' are defined)

# --- DYNAMIC POLICY CHECK ---
IS_ORDERED_POLICY = "Ordered" in agent.policy.__class__.__name__

print(f"Policy detected: {agent.policy.__class__.__name__}")
print(f"Gemini Injection Mode: {'SORTED (Ordered Trajectory)' if IS_ORDERED_POLICY else 'UNSORTED (Permutation Invariant)'}")

# ==========================================
# 1. MULTI-AGENT DEBATE CONFIGURATION
# ==========================================
PROJECT_ID = "crn-evolution"

gemini_task_desc = (
    f"Implement a Chemical Reaction Network that achieves Robust Perfect Adaptation (RPA) via Integral Control. "
    f"The system has 2 inputs: u1 (Reference Setpoint) and u2 (Disturbance). "
    f"Goal 1 (Tracking): The output species 'r' must converge exactly to the concentration of u1 at steady state. "
    f"Goal 2 (Robustness): The output must remain at u1 regardless of the value of u2 (the disturbance). "
    f"The CRN must use species {species_labels} and start from this template: {crn_template}. "
    f"Select exactly {max_added_reactions} reactions. "
    f"Target steady-state tracking error < 0.001. "
    f"Note that in this task you must use EXACTLY {max_added_reactions} reactions from the library provided."
    f"The available species are {species_labels} and solely them."
)

# Instantiate the Debate System
# Narrator/Player use Pro (Smart), others use Flash (Fast) for cost/speed efficiency
debate_system = VertexMultiAgentDebate(
    project_id=PROJECT_ID, 
    location="global", 
    fast_model_name="gemini-2.5-flash",        # Opportunist, Contrarian, Skeptic
    smart_model_name="gemini-3-pro-preview",   # Narrator, Player
    track_top_k=50
)

gemini_schedule = 10  # Run the debate every 10 epochs

# ==========================================
# 2. TRAINING LOOP
# ==========================================
if train_flag:
    agent.policy.train()
    
    # Optional: Force a "warm start" debate at epoch 1 to populate HoF early
    warm_start_triggered = False

    debate_transcript_file = f"debate_trainscript_log_{task_name}_{time.strftime('%Y%m%d_%H%M%S')}.txt"

    for i in tqdm(range(epoch_num)):
        
        # --- A. Standard RL Step ---
        mult_env.reset()
        for j in range(max_added_reactions):
            observations = mult_env.observe(observer, tensorizer)
            actions, raw_actions = agent.act(observations, actuator)
            out = mult_env.step(actions, stepper, raw_actions=raw_actions)
        
        rewards = mult_env.get_reward(compute_reward)
        
        # Add current batch to Hall of Fame
        mult_env.hall_of_fame.add_all(mult_env.envs)

        # Logging Standard Metrics
        successful_count = sum(1 for env in mult_env.envs if not env.state.last_task_info.get('has_diverged', False))
        if logger:
            logger.log_metric("Successful Environments (%)", successful_count/N, step=i)
        
        
        # --- B. GEMINI MULTI-AGENT DEBATE PHASE ---
        # Trigger if schedule matches OR if it's the first epoch (to seed the HoF)
        should_run_debate = (i > 0 and i % gemini_schedule == 0) or (i == 1 and not warm_start_triggered)

        should_run_debate = (i > 0 and i % gemini_schedule == 0) or (i == 1 and not warm_start_triggered)

        if should_run_debate:
            if i == 1: warm_start_triggered = True
            
            start_time = time.time()
            print(f"\n[Gemini] Epoch {i}: Initiating Multi-Agent Debate...")
            
            # 1. Run the Debate -> NOW RETURNS 'STORY'
            candidates, story = debate_system.run_debate_and_generate(
                hall_of_fame_iter=mult_env.hall_of_fame,
                task_description=gemini_task_desc,
                reaction_library=library,
                iteration=i,
                max_added_reactions=max_added_reactions
            )
            
            # 2. SAVE THE STORY TO FILE
            with open(debate_transcript_file, "a", encoding="utf-8") as f:
                f.write(story)
            print(f"[Gemini] Story saved to {debate_transcript_file}, for epoch {i}.")
            
            # 2. Evaluate, Transplant & Track
            # This simulates the "Player's" JSON suggestions in the real physics engine
            new_gemini_envs = debate_system.evaluate_and_transplant(
                candidates=candidates,
                crn_template=crn_template,
                max_added_reactions=max_added_reactions,
                library=library,
                stepper=stepper,
                actuator=actuator,
                compute_reward_func=compute_reward,
                is_ordered_policy=IS_ORDERED_POLICY,
                logger=logger
            )

            # 3. Inject into Agent's Hall of Fame
            if new_gemini_envs:
                mult_env.hall_of_fame.add_all(new_gemini_envs)

            # Logging & Timing
            elapsed_time = time.time() - start_time
            print(f"[Gemini] Debate Finished in {elapsed_time:.2f}s. {len(new_gemini_envs)} valid candidates added.")
            
            if logger:
                logger.log_metric("Gemini Candidates (Debate)", len(new_gemini_envs), step=i)
                logger.log_metric("Timing: Debate Duration (s)", elapsed_time, step=i)
                
                # Log HoF stats to see if the Debate Agents are beating the RL Agents
                if len(mult_env.hall_of_fame) > 0:
                    best_env = mult_env.hall_of_fame[0] # Assumes HoF is sorted
                    worst_env = mult_env.hall_of_fame[len(mult_env.hall_of_fame)-1]
                    logger.log_metric("HoF Best Loss", best_env.state.last_task_info.get('reward'), step=i)


        # --- C. Agent Update (Self-Imitation / PPO) ---
        # The agent now trains on its own experience PLUS the "Synthesized Dreams" of the Debate System
        agent.update(
            rewards, 
            step_iteration=i, 
            hof=mult_env.hall_of_fame, 
            observer=observer, 
            tensorizer=tensorizer, 
            stepper=stepper, 
            use_sil=True, 
            sil_weighting_scheme='uniform', 
            sil_batch_size=None
        )

        if i % render_schedule == 0:
            mult_env.render(rewards, n_best=render_n_best, disregarded_percentage=render_disregard_percentage, mode=render_mode)

# %%
hall_of_fame_crns = [env.state for env in mult_env.hall_of_fame]
if save_flag:
    if not os.path.exists('models'):
        os.makedirs('models')
    if not os.path.exists('hof'):
        os.makedirs('hof')
    torch.save(agent.policy.state_dict(), 'models/' + save_filename)
    torch.save(hall_of_fame_crns, 'hof/hall_of_fame_' + save_filename)


