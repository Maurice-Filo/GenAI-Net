# %%
# Set up the file path
import sys
import os
parent_dir = os.path.abspath(os.path.join(os.getcwd(), '..', '..'))
sys.path.append(parent_dir)
print('Working directory set to:', parent_dir)
task_name = "Comparator_5s_6r_MAK_REINFORCE"

# %%
# Import packages
# Import general packages
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import torch
from matplotlib import pyplot as plt
from pytorch_lightning.loggers import CometLogger
import numpy as np
from itertools import product
from tqdm import tqdm

# Import Agent-Environment packages
from RL4CRN.environments.environment import Environment
from RL4CRN.environments.parallel_environments import ParallelEnvironments
from RL4CRN.environments.serial_environments import SerialEnvironments
from RL4CRN.agents.reinforce_agent import REINFORCEAgent
from RL4CRN.policies.add_reaction_by_ordered_index import AddReactionByOrderedIndex
from RL4CRN.policies.add_reaction_by_index import AddReactionByIndex

# Import Interface packages
from RL4CRN.env2agent_interface.explicit_observer import ExplicitObserver
from RL4CRN.env2agent_interface.explicit_tensorizer import ExplicitTensorizer
from RL4CRN.agent2env_interface.library_actuator import LibraryActuator
from RL4CRN.agent2env_interface.iocrn_stepper import IOCRNStepper

# Import CRN packages
from RL4CRN.iocrns.iocrn import IOCRN
from RL4CRN.iocrns.reactions import MassAction
from RL4CRN.utils.ic import IC
from RL4CRN.iocrns.reaction_library import construct_mass_action_library

# Import Reward packages
from RL4CRN.rewards.deterministic import track_relationship

# %%
# Set the logger to use Comet
timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
api_key = "vhIR3uyqsKyU4L7SA8fLCfTSC"
logger = CometLogger(
    api_key=api_key,
    project=f"{task_name}",        
    workspace="redsnic", 
    name=f'{task_name}_{timestamp}',
)
logger = logger.experiment

# %%
# Construct the template CRN
# nothing to A with rate u1
r1 = MassAction(reactant_labels=[], product_labels=['A'], input_channels=['u_1'], params=[1.], params_controllability=[True])
# nothing to B with rate u2
r2 = MassAction(reactant_labels=[], product_labels=['B'], input_channels=['u_2'], params=[1.], params_controllability=[True])
# degradation of Delta with rate u3
r3 = MassAction(reactant_labels=['D'], product_labels=[], input_channels=['u_3'], params=[1.], params_controllability=[True])

# Update species list for the template context
species_labels = ['A', 'B', 'Z1', 'Z2', 'D']

# We define the output species. You didn't specify one, so I will assume 'Z' acts as the output based on previous context, 
# but you can change 'Z' to 'Delta' or 'A' if needed.
crn_template = IOCRN([r1, r2, r3], output_labels=['D']) 
crn_template.compile()

p = crn_template.num_inputs # Number of inputs of the IOCRNs
print("Template CRN:")
print(crn_template)

# Construct the library of possible reactions
# The library will now generate interactions between A, B, Z, and Delta
library = construct_mass_action_library(species_labels=species_labels, order=2)
# library.remove_reactions([
#     5, 6, 7, 8, 9, 10, 11, 15, 16, 17, 18, 19, 20, 23, 24, 25, 26, 27, 28, 
#     29, 30, 31, 32, 33, 34, 35, 36, 37, 40, 41, 42, 44, 45, 47, 48, 49, 50, 51, 
#     52, 53, 58, 59, 61, 62, 63, 64, 65, 66, 67, 71, 72, 73, 74, 75, 76, 77, 78, 
#     79, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 
#     98, 99, 101, 102, 103, 104, 105, 107, 108, 109, 110, 111, 112, 113, 115, 
#     116, 117, 118, 119, 121, 122, 123, 124, 125, 126, 127, 128, 129, 130, 131, 
#     132, 133, 134, 135, 136, 137, 138, 139, 140, 141, 142, 144, 145, 146, 147, 
#     148, 149, 150, 152, 153, 154, 155, 156, 158, 159, 160, 161, 162, 163, 164, 
#     166, 167, 168, 170, 171, 174, 175, 176, 177, 178, 179, 180, 184, 185, 188, 
#     189, 190, 191, 192, 193, 194, 198, 199, 202, 203, 204, 205, 206, 207, 208
# ])
library.remove_reactions([ 
    6, 7, 8, 9, 10, 11, 12, 13, 14, 
    21, 22, 23, 24, 25, 26, 27, 
    31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 
    55, 56, 57, 58, 59, 60, 
    62, 63, 66, 67, 68, 69, 70, 71, 72, 73, 74, 
    82, 83, 86, 87, 88, 89, 90, 91, 92, 93, 94, 
    102, 103, 106, 107, 108, 109, 110, 111, 112, 113, 114, 
    121, 122, 123, 124, 125, 126, 127, 128, 129, 130, 131, 132, 133, 134, 135, 136, 137, 138, 139, 140, 
    141, 142, 143, 144, 145, 146, 147, 148, 149, 150, 151, 152, 153, 154, 155, 156, 157, 158, 159, 160, 
    161, 163, 164, 165, 166, 167, 168, 171, 172, 173, 174, 175, 176, 177, 178, 179, 180, 
    181, 183, 184, 185, 186, 187, 188, 191, 192, 193, 194, 195, 196, 197, 198, 199, 200, 
    201, 203, 204, 205, 206, 207, 208, 211, 212, 213, 214, 215, 216, 217, 218, 219, 220, 
    221, 222, 223, 224, 225, 226, 227, 228, 229, 230, 231, 232, 233, 234, 235, 236, 237, 238, 239, 240, 
    241, 242, 244, 245, 246, 247, 248, 249, 250, 251, 252, 255, 256, 257, 258, 259, 260, 
    261, 262, 264, 265, 266, 267, 268, 269, 270, 271, 272, 275, 276, 277, 278, 279, 280, 
    281, 282, 284, 285, 286, 287, 288, 289, 290, 291, 292, 295, 296, 297, 298, 299, 300, 
    302, 303, 307, 308, 309, 310, 311, 312, 313, 314, 315, 
    322, 323, 327, 328, 329, 330, 331, 332, 333, 334, 335, 
    342, 343, 347, 348, 349, 350, 351, 352, 353, 354, 355, 
    362, 363, 367, 368, 369, 370, 371, 372, 373, 374, 375, 
    382, 383, 387, 388, 389, 390, 391, 392, 393, 394, 395, 
    402, 403, 407, 408, 409, 410, 411, 412, 413, 414, 415
])
crn_template.set_library_context(library)

M = len(library.reactions) # Number of possible reactions

K = library.get_num_parameters() # Total number of parameters in all the reactions of the library
print("Library of possible reactions:")
print(library)
print("------------------------------------------------")

# %%
# Device
device = 'cuda' if torch.cuda.is_available() else 'cpu'
print(f'Using device: {device}')
print(f'Number of CPUs available: {os.cpu_count()}') 

# %%
# Flags and filenames
save_flag = True                                                # Save the agent checkpoint
load_flag = False                                               # Load the agent checkpoint 
train_flag = True                                               # Train the agent
save_sheet_flag = True                                          # Save the configuration to an Excel sheet

save_filename = timestamp + '.pth'                              # Filename for saving the agent checkpoint
load_filename = ''                                              # Filename for loading the agent checkpoint
file_name = f"{task_name}.xlsx"             # Filename for saving the Excel sheet

# %%
# Hyperparameters
max_added_reactions = 8                             # Maximum number of reactions
N_CPUs = os.cpu_count()                             # Number of CPUs          
N = 10*N_CPUs                                       # Number of samples (batch size)    
width = 1024                                        # Width of the neural networks  
depth = 5                                           # Depth of the neural networks 
deep_layer_size = 1024*10                           # Size of the deep layer encoding the CRNs
allow_input_influence = False                       # Allow input influence in the policy
learning_rate = 1e-4                                # Learning rate for the optimizer 
hall_of_fame_size = 30                              # Size of the hall of fame  
entropy_scheduler = {                               # Entropy scheduler parameters 
    'entropy_weight': 1e-3, 
    'topk_entropy_weight' : 1.0,
    'remainder_entropy_weight' : 1.0,
    'entropy_update_coefficient': 1, 
    'entropy_schedule': 1000, 
    'minimum_entropy_weight': 0.0
}
entropy_weights_per_head = {'structure': 2.0, 'continuous': 1.0, 'discrete': 0.0, 'input_influence': 0.0} 
structure_head_temperature = {"target_entropy_ratio_to_max": np.log(5)/np.log(M), "initial_temperature": 1.0, "rate": 0.0, "current_temperature": 1.0}
risk_scheduler = {                                  # Risk scheduler parameters
    'risk': 0.9, 
    'risk_update': 0.0, 
    'max_risk': 1.0, 
    'risk_schedule': 1000
}
epoch_num = 300                                     # Number of epochs for training
render_schedule = 10                                 # Render every # of epochs
render_mode = {                                     # Mode of the experiment
    'style': 'logger', 
    'task': 'transients', 
    'format': 'image',
    'topology': True,
    'bounds': [2.5]
}
# Ordering specific parameters
ordering_parameters = {
    'enforce_ordering': False,
    'constraint_weight' : float('inf')
}
# SIL settings
sil_settings = {
    'sil_loss_weight': 1.0,
    'sil_use_adaptive_baseline': False,
    'sil_baseline_annealing_rate': 0.95
}
render_n_best = 10                                      # Number of best CRNs to plot responses for
render_disregard_percentage = risk_scheduler['risk']    # Percentage of worst CRNs to disregard in the responses plotting

# Parameter distribution for the reactions added by the agent
continuous_distribution = {"type": 'lognormal_1D'}        

# Time horizon for the simulation
t_f = 100                                               # Final time for the simulation
N_t = 1000                                              # Number of time steps in the simulation
time_horizon = np.linspace(0, t_f, N_t, dtype=np.float32)

# Construct the IOCRN inputs
b_vals = [0.5, 1.0, 1.5]
a_vals = [1., 2., 3.]
perturbances = [0.5, 1.0, 1.5]  # possible perturbance
u_list = [np.array(u) for u in product(a_vals, b_vals, perturbances)] # list of input combinations, each input is a numpy array of shape (p,)

# Note: here we don't have a setpoint, we dynamically want D = max(0, A - B)

# Construct the IOCRN initial conditions
ic = IC(names=species_labels, values=[[0.01, 0.01, 0.01, 0.01, 0.01]]) # non-zero initial conditions for all species

# Construct the weights for the performance metric
w = np.ones(N_t)
w[(len(w)//5)*4:] = w[(len(w)//5)*4:]*2
w[:(len(w)//5)] = w[:(len(w)//5)]*0.25
w = w[np.newaxis, :]

# Construct the compute reward routine
def compute_reward(state):
    x0_list = ic.get_ic(state)
    return track_relationship(state, u_list, x0_list, time_horizon, w, species_names=["A", "B", "D"], relationship_func=lambda A, B, D: np.maximum(0, A - B) - D, norm=1, LARGE_NUMBER=1e4)

# %%
# Log the code and hyperparameters
# Log the code of the current file
current_file_path = os.path.abspath(__file__)
logger.log_code(file_name=os.path.basename(current_file_path))

# Log the hyperparameters
hyperparameters = {
    'species_labels': str(species_labels),
    'crn_template': str(crn_template),
    'library': str(library),
    'max_added_reactions': max_added_reactions,
    'N_CPUs': N_CPUs,
    'N': N,
    'width': width,
    'depth': depth,
    'deep_layer_size': deep_layer_size,
    'allow_input_influence': allow_input_influence,
    'learning_rate': learning_rate,
    'hall_of_fame_size': hall_of_fame_size,
    'entropy_scheduler': entropy_scheduler,
    'structure_head_temperature': structure_head_temperature,
    'risk_scheduler': risk_scheduler,
    'epoch_num': epoch_num,
    'render_schedule': render_schedule,
    'render_mode': render_mode,
    'render_n_best': render_n_best,
    'render_disregard_percentage': render_disregard_percentage,
    'continuous_distribution': continuous_distribution,
    'tf': t_f,
    'N_t': N_t,
    'u_list': [ u.tolist() for u in u_list ],
    'ic': str(ic),
    'w': w.tolist()
}
logger.log_parameters(hyperparameters)

# %%
# Save the experiment configuration to an Excel sheet
if save_sheet_flag:
    sheet_name = "Data"
    headers = [
        "Timestamp", "URL",
        "Successful", "Epochs Completed", "Saved", "Comments",
        "Entropy Scheduler",
        "Learning Rate", "Epochs #",
        "(m, n, p, N)",
        "NN Depth", "NN Width", "Deep Layer Size", "CPUs #",
        "Risk Scheduler",
        "Render Schedule", "HoF Size",
        "Simulation Time", "Time Steps #",
        "Initial Conditions #", "Input Scenarios#",
        "Continuous Distribution", 
        "Structure Head Temperature"
    ]

    data_row = [
        timestamp, logger.url,
        None, None, None, None,
        str(entropy_scheduler),
        str(learning_rate), epoch_num,
        str((max_added_reactions, len(species_labels), p, N)),
        depth, width, deep_layer_size, N_CPUs,
        str(risk_scheduler),
        render_schedule, hall_of_fame_size,
        t_f, N_t, len(ic.values), len(u_list),
        str(continuous_distribution), 
        str(structure_head_temperature)
    ]

    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    # Write headers if sheet is empty
    if ws.max_row == 1 and ws.max_column == 1 and ws.cell(row=1, column=1).value is None:
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

    # Append experiment as next row
    next_row = ws.max_row + 1
    for col, value in enumerate(data_row, start=1):
        ws.cell(row=next_row, column=col, value=value)

    # Freeze header row and add filter
    ws.freeze_panes = "B1" 
    ws.auto_filter.ref = ws.dimensions

    # === Auto-fit column widths (except URL column) ===
    # URL column is column 2 (B), we leave its width unchanged.
    url_col_index = 2

    for col in range(1, ws.max_column + 1):
        if col == url_col_index:
            continue  # keep URL column width as-is

        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            value = cell.value
            if value is not None:
                # Convert to string to measure length
                length = len(str(value))
                if length > max_length:
                    max_length = length

        # Some padding so text isn't touching the cell border
        adjusted_width = max_length + 1 if max_length > 0 else 10
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(file_name)
    print(f"New experiment data saved in row {next_row} of '{file_name}'.")

# %%
# Construct parallel environments
crn_0 = crn_template.clone()
mult_env = ParallelEnvironments([Environment(crn_0, max_added_reactions, logger=logger, logger_schedule=1) for _ in range(N)], hall_of_fame_size=hall_of_fame_size, N_CPUs=N_CPUs, logger=logger)
# mult_env = SerialEnvironments([Environment(crn_0, max_added_reactions, logger=logger, logger_schedule=1) for _ in range(N)], hall_of_fame_size=hall_of_fame_size, logger=logger)

# %%
# Construct the policy
encoder_attributes = {"hidden_size": width, "num_layers": depth}
structure_head_attributes = {"hidden_size": width, "num_layers": depth}
rate_head_attributes = {"hidden_size": width, "num_layers": depth}
input_influence_head_attributes = {"hidden_size": width, "num_layers": depth}
masks = {"continuous": library.get_parameter_mask(mode="continuous"), "discrete": library.get_parameter_mask(mode="discrete"), "logit": library.get_logit_mask()}
policy = AddReactionByOrderedIndex(M, K, p, encoder_attributes, deep_layer_size, structure_head_attributes, rate_head_attributes, input_influence_head_attributes, target_set_size=crn_template.num_reactions+max_added_reactions, allow_input_influence=False, masks=masks, device=device, continuous_distribution=continuous_distribution, entropy_weights_per_head=entropy_weights_per_head,
                                    combinatorial_bias_enabled=ordering_parameters["enforce_ordering"], constraint_strength=ordering_parameters["constraint_weight"])

if ordering_parameters["enforce_ordering"]:
    policy = AddReactionByOrderedIndex(M, K, p, encoder_attributes, deep_layer_size, structure_head_attributes, rate_head_attributes, input_influence_head_attributes, target_set_size=crn_template.num_reactions+max_added_reactions, allow_input_influence=False, masks=masks, device=device, continuous_distribution=continuous_distribution, entropy_weights_per_head=entropy_weights_per_head,
                                        combinatorial_bias_enabled=ordering_parameters["enforce_ordering"], constraint_strength=ordering_parameters["constraint_weight"])
else:
    policy = AddReactionByIndex(M, K, p, encoder_attributes, deep_layer_size, structure_head_attributes, rate_head_attributes, input_influence_head_attributes, allow_input_influence=False, masks=masks, device=device, continuous_distribution=continuous_distribution, entropy_weights_per_head=entropy_weights_per_head)

# Construct the agent
agent = REINFORCEAgent(policy, allow_input_influence=False, logger=logger, learning_rate=learning_rate, entropy_scheduler=entropy_scheduler, risk_scheduler=risk_scheduler, sil_settings=sil_settings, device=device)
if load_flag:
    agent.policy.load_state_dict(torch.load(load_filename+'.pth', map_location=device))

# %%
# Construct the interfaces
observer = ExplicitObserver(reaction_library=library, allow_input_observation=allow_input_influence)
tensorizer = ExplicitTensorizer(device=device)
actuator = LibraryActuator(reaction_library=library)
stepper = IOCRNStepper()

# %%
# Training Loop   
if train_flag:
    agent.policy.train()
    for i in tqdm(range(epoch_num)):
        mult_env.reset()
        for j in range(max_added_reactions):
            observations = mult_env.observe(observer, tensorizer)
            actions, raw_actions = agent.act(observations, actuator)
            out = mult_env.step(actions, stepper, raw_actions=raw_actions)
        rewards = mult_env.get_reward(compute_reward)
        agent.update(rewards, step_iteration=i, hof=mult_env.hall_of_fame, observer=observer, tensorizer=tensorizer, stepper=stepper, use_sil=True, sil_weighting_scheme='uniform', sil_batch_size=None)
        if i % render_schedule == 0:
            mult_env.render(rewards, n_best=render_n_best, disregarded_percentage=render_disregard_percentage, mode=render_mode)

# %%
# Save the agent and the hall of fame CRNs
hall_of_fame_crns = [env.state for env in mult_env.hall_of_fame]
if save_flag:
    if not os.path.exists('models'):
        os.makedirs('models')
    if not os.path.exists('hof'):
        os.makedirs('hof')
    torch.save(agent.policy.state_dict(), 'models/' + save_filename)
    torch.save(hall_of_fame_crns, 'hof/hall_of_fame_' + save_filename)


