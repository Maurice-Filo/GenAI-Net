# %%
# Set up the file path
import sys
import os
parent_dir = os.path.abspath(os.path.join(os.getcwd(), '..', '..'))
sys.path.append(parent_dir)
print('Working directory set to:', parent_dir)

# %%
# Import general packages
from openpyxl import Workbook, load_workbook
from datetime import datetime
import torch
from matplotlib import pyplot as plt
from pytorch_lightning.loggers import CometLogger
import numpy as np
from itertools import product
from tqdm import tqdm

# Import Agent-Environment packages
from RL4CRN.environments.environment import Environment
from RL4CRN.environments.parallel_environments import ParallelEnvironments
from RL4CRN.environments.serial_environments import SerialEnvironments
from RL4CRN.agents.reinforce_agent import REINFORCEAgent
from RL4CRN.policies.add_reaction_by_ordered_index import AddReactionByOrderedIndex

# Import Interface packages
from RL4CRN.env2agent_interface.explicit_observer import ExplicitObserver
from RL4CRN.env2agent_interface.explicit_tensorizer import ExplicitTensorizer
from RL4CRN.agent2env_interface.library_actuator import LibraryActuator
from RL4CRN.agent2env_interface.iocrn_stepper import IOCRNStepper

# Import CRN packages
from RL4CRN.iocrns.iocrn import IOCRN
from RL4CRN.iocrns.reactions import MassAction
from RL4CRN.utils.ic import IC
from RL4CRN.iocrns.reaction_library import construct_mass_action_library

# Import Reward packages
from RL4CRN.rewards.deterministic import dynamic_tracking_error

# %%
# Set the logger to use Comet
timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
api_key = "o77J6VCMDamustkfJuMXZ2jdV"
logger = CometLogger(
    api_key=api_key,
    project="multistability_2species_5rxns_MAK_REINFORCE",        
    workspace="maurice-filo", 
    name=f'multistability_2species_5rxns_MAK_REINFORCE_{timestamp}',
)
logger = logger.experiment

# %%
# Construct the template CRN
r1 = MassAction(reactant_labels=['X_1'], product_labels=[], input_channels=['u_1'], params=[1.0], params_controllability=[True])
r2 = MassAction(reactant_labels=['X_2'], product_labels=[], input_channels=[None], params=[1.0], params_controllability=[True])
crn_template = IOCRN([r1, r2], output_labels=['X_1', 'X_2'])
crn_template.compile()
p = crn_template.num_inputs # Number of inputs of the IOCRNs
print("Template CRN:")
print(crn_template)

# Construct the library of possible reactions
species_labels = ['X_1', 'X_2']
library = construct_mass_action_library(species_labels=species_labels, order_reactants=2, order_products=2)
crn_template.set_library_context(library)
M = len(library.reactions) # Number of possible reactions
K = library.get_num_parameters() # Total number of parameters in all the reactions of the library
print("Library of possible reactions:")
print(library)
print("------------------------------------------------")

# %%
# Device
device = 'cuda' if torch.cuda.is_available() else 'cpu'
print(f'Using device: {device}')
print(f'Number of CPUs available: {os.cpu_count()}') 

# %%
# Flags and filenames
save_flag = True                                                # Save the agent checkpoint
load_flag = False                                              # Load the agent checkpoint 
train_flag = True                                               # Train the agent
save_sheet_flag = True                                          # Save the configuration to an Excel sheet

save_filename = timestamp + '.pth'                              # Filename for saving the agent checkpoint
load_filename = ''                                              # Filename for loading the agent checkpoint
file_name = "multistability_2species_5rxns_MAK_REINFORCE.xlsx"                           # Filename for saving the Excel sheet

# %%
# Hyperparameters
max_added_reactions = 5                             # Maximum number of reactions
N_CPUs = os.cpu_count()                             # Number of CPUs          
N = 10*N_CPUs                                       # Number of samples (batch size)    
width = 1024                                        # Width of the neural networks  
depth = 5                                           # Depth of the neural networks 
deep_layer_size = 1024*10                           # Size of the deep layer encoding the CRNs
allow_input_influence = False                       # Allow input influence in the policy
learning_rate = 1e-4                                # Learning rate for the optimizer 
hall_of_fame_size = 10                              # Size of the hall of fame  
entropy_scheduler = {                               # Entropy scheduler parameters 
    'entropy_weight': 1e-2, 
    'entropy_update_coefficient': 0.1, 
    'entropy_schedule': 1000, 
    'minimum_entropy_weight': 0.0
}
entropy_weights_per_head = {'structure': 1.0, 'continuous': 1.0, 'discrete': 0.0, 'input_influence': 0.0}
structure_head_temperature = {"target_entropy_ratio_to_max": np.log(3)/np.log(M), "initial_temperature": 1.0, "rate": 0.0, "current_temperature": 1.0}
risk_scheduler = {                                  # Risk scheduler parameters
    'risk': 0.9, 
    'risk_update': 0.0, 
    'max_risk': 1.0, 
    'risk_schedule': 1000
}
epoch_num = 300                                     # Number of epochs for training
render_schedule = 10                                 # Render every # of epochs
render_mode = {                                            # Mode of the experiment
    'style': 'logger', 
    'task': 'phase_plot',
    'format': 'image',
    'topology': True,
    'bounds': [1.2, 1.2]
}
render_n_best = 3                                   # Number of best CRNs to plot responses for
render_disregard_percentage = risk_scheduler['risk']                   # Percentage of worst CRNs to disregard in the responses plotting

# Parameter distribution for the reactions added by the agent
continuous_distribution = {'type': 'lognormal_1D'}

# Time horizon for the simulation
t_f = 100                                           # Final time for the simulation
N_t = 1000                                          # Number of time steps
time_horizon = np.linspace(0, t_f, N_t, dtype=np.float32)

# Construct the IOCRN inputs
u_list = [np.array([1], dtype=np.float32)]  # list of input combinations, each input is a numpy array of shape (p,)

# Construct the IOCRN initial conditions
ic_values_diagonal_1 = [[x/10, x/10 - 0.1] for x in range(1, 11)]
ic_values_diagonal_2 = [[x/10, x/10 + 0.1] for x in range(0, 10)]
ic_values_cluster_1 = [ [.9, 0], [1, 0], [1.1, 0], 
                        [.9, .1], [1, .1], [1.1, .1]]
ic_values_cluster_2 = [ [0, .9], [0, 1], [0, 1.1], 
                        [.1, .9], [.1, 1], [.1, 1.1]]
ic_values = ic_values_diagonal_1 + ic_values_diagonal_2 + ic_values_cluster_1 + ic_values_cluster_2
ic = IC(names=species_labels, values=ic_values)

# Construct the desired fixed points
r_list = [np.array([1, 0])] * len(ic_values_diagonal_1) + \
         [np.array([0, 1])] * len(ic_values_diagonal_2) + \
         [np.array([1, 0])] * len(ic_values_cluster_1) + \
         [np.array([0, 1])] * len(ic_values_cluster_2)

# Construct the weights for the performance metric
w = np.ones(N_t)
w[(len(w)//5)*4:] = w[(len(w)//5)*4:]*2
w[:(len(w)//5)] = w[:(len(w)//5)]*0.25
w = w[np.newaxis, :]

# Construct the compute reward routine
def compute_reward(state):
    x0_list = ic.get_ic(state)
    return dynamic_tracking_error(state, u_list, x0_list, time_horizon, r_list, w, norm=1, relative=False, LARGE_NUMBER=1e4)

# %%
# Save the experiment discription in an Excel sheet
if save_sheet_flag:
    sheet_name = "Data"
    headers = ["Timestamp", "URL", 
               "Maximum Number of Added Reactions", 
               "Number of Species", 
               "Number of Inputs", 
               "Batch Size", 
               "Allow Input Influence", 
               "Learning Rate", "Number of Epochs", "Neural Network Depth", "Neural Network Width", "Deep Layer Size", "Number of CPUs",
               "Entropy Scheduler", 
               "Risk Scheduler",
               "Render Schedule", "Hall of Fame Size", 
               "Simulation Time", "Number of Time Steps", "Initial Condition Scenarios", "Input Scenarios",
               "Continuous Distribution", "Entropy Weights per Head", "Structure Head Temperature",
               "Epochs Completed", "Successful", "Saved", "Useful", "Comments"]
    data_row = [timestamp, logger.url, 
                max_added_reactions, 
                len(species_labels), 
                p, N,
                allow_input_influence, 
                learning_rate, epoch_num, depth, width, deep_layer_size, N_CPUs, 
                str(entropy_scheduler),
                str(risk_scheduler),
                render_schedule, hall_of_fame_size,
                t_f, N_t, len(ic.values), len(u_list),
                str(continuous_distribution), str(entropy_weights_per_head), str(structure_head_temperature),
                None, None, None, None]

    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    if ws.max_column < 2:
        for i, header in enumerate(headers, start=1):
            ws.cell(row=i, column=1, value=header)
    next_col = ws.max_column + 1
    for i, value in enumerate(data_row, start=1):
        ws.cell(row=i, column=next_col, value=value)

    wb.save(file_name)
    print(f"New experiment data saved in column {next_col} of '{file_name}'.")

# %%
# Construct parallel environments
crn_0 = crn_template.clone()
mult_env = ParallelEnvironments([Environment(crn_0, max_added_reactions, logger=logger, logger_schedule=1) for _ in range(N)], hall_of_fame_size=hall_of_fame_size, N_CPUs=N_CPUs, logger=logger)
# mult_env = SerialEnvironments([Environment(crn_0, max_added_reactions, logger=logger, logger_schedule=1) for _ in range(N)], hall_of_fame_size=hall_of_fame_size, logger=logger)

# %%
# Construct the policy
encoder_attributes = {"hidden_size": width, "num_layers": depth}
structure_head_attributes = {"hidden_size": width, "num_layers": depth}
rate_head_attributes = {"hidden_size": width, "num_layers": depth}
input_influence_head_attributes = {"hidden_size": width, "num_layers": depth}
masks = {"continuous": library.get_parameter_mask(mode="continuous"), "discrete": library.get_parameter_mask(mode="discrete"), "logit": library.get_logit_mask()}
policy = AddReactionByOrderedIndex(M, K, p, encoder_attributes, deep_layer_size, structure_head_attributes, rate_head_attributes, input_influence_head_attributes, target_set_size=crn_template.num_reactions+max_added_reactions, allow_input_influence=False, masks=masks, device=device, continuous_distribution=continuous_distribution, entropy_weights_per_head=entropy_weights_per_head, combinatorial_bias_enabled=True)
# Construct the agent
agent = REINFORCEAgent(policy, allow_input_influence=False, logger=logger, learning_rate=learning_rate, entropy_scheduler=entropy_scheduler, risk_scheduler=risk_scheduler, device=device)
if load_flag:
    agent.policy.load_state_dict(torch.load(load_filename+'.pth', map_location=device))

# %%
# Construct the interfaces
observer = ExplicitObserver(reaction_library=library, allow_input_observation=allow_input_influence)
tensorizer = ExplicitTensorizer(device=device)
actuator = LibraryActuator(reaction_library=library)
stepper = IOCRNStepper()

# %%
# Training Loop     
if train_flag:
    agent.policy.train()
    for i in tqdm(range(epoch_num)):
        mult_env.reset()
        for j in range(max_added_reactions):
            observations = mult_env.observe(observer, tensorizer)
            actions = agent.act(observations, actuator)
            out = mult_env.step(actions, stepper)
        rewards = mult_env.get_reward(compute_reward)
        agent.update(rewards, step_iteration=i)
        if i % render_schedule == 0:
            mult_env.render(rewards, n_best=render_n_best, disregarded_percentage=render_disregard_percentage, mode=render_mode)

# %%
# Test the model
agent.policy.eval()
mult_env.reset()
for j in range(max_added_reactions):
    observations = mult_env.observe(observer, tensorizer)
    actions = agent.act(observations, actuator)
    out = mult_env.step(actions, stepper)
rewards = mult_env.get_reward(compute_reward)

# Gather the CRNs from the environments
crns = mult_env.gather()

# Sort the CRNs by rewards
sorted_crns_rewards = sorted(zip(crns, rewards), key=lambda x: x[1])

# %%
# Plot the results versus time
n_plot = 10
ax = None
for i in range(n_plot):
    x0_list = ic.get_ic(sorted_crns_rewards[i][0])
    time_horizon, x_list, y_list, last_task_info = sorted_crns_rewards[i][0].transient_response(u_list, x0_list, time_horizon)
    fig, ax = sorted_crns_rewards[i][0].plot_transient_response(axes=ax)
    print(f"IOCRN {i}, Reward: {sorted_crns_rewards[i][1]}")
    print(sorted_crns_rewards[0][0])

# %%
# Plot the results versus time
n_plot = 10
ax = None
for i in range(n_plot):
    x0_list = ic.get_ic(sorted_crns_rewards[i][0])
    time_horizon, x_list, y_list, last_task_info = sorted_crns_rewards[i][0].transient_response(u_list, x0_list, time_horizon)
    fig, ax = sorted_crns_rewards[i][0].plot_phase_portrait(axis=ax)
    print(f"IOCRN {i}, Reward: {sorted_crns_rewards[i][1]}")
    print(sorted_crns_rewards[0][0])

# %%
hall_of_fame_crns = [env.state for env in mult_env.hall_of_fame]
if save_flag:
    torch.save(agent.policy.state_dict(), save_filename)
    torch.save(hall_of_fame_crns, 'hall_of_fame_' + save_filename)


