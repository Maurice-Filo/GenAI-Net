# %%
# Set up the file path
import sys
import os
parent_dir = os.path.abspath(os.path.join(os.getcwd(), '..', '..'))
sys.path.append(parent_dir)
print('Working directory set to:', parent_dir)

# %%
# Import packages
# Import general packages
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import torch
from matplotlib import pyplot as plt
from pytorch_lightning.loggers import CometLogger
import numpy as np
from itertools import product
from tqdm import tqdm

# Import Agent-Environment packages
from RL4CRN.environments.environment import Environment
from RL4CRN.environments.parallel_environments import ParallelEnvironments
from RL4CRN.environments.serial_environments import SerialEnvironments
from RL4CRN.agents.reinforce_agent import REINFORCEAgent
from RL4CRN.policies.add_reaction_by_ordered_index import AddReactionByOrderedIndex
from RL4CRN.policies.add_reaction_by_index import AddReactionByIndex

# Import Interface packages
from RL4CRN.env2agent_interface.explicit_observer import ExplicitObserver
from RL4CRN.env2agent_interface.explicit_tensorizer import ExplicitTensorizer
from RL4CRN.agent2env_interface.library_actuator import LibraryActuator
from RL4CRN.agent2env_interface.iocrn_stepper import IOCRNStepper

# Import CRN packages
from RL4CRN.iocrns.iocrn import IOCRN
from RL4CRN.iocrns.reactions import MassAction
from RL4CRN.utils.ic import IC
from RL4CRN.iocrns.reaction_library import construct_mass_action_library

# Import Reward packages
from RL4CRN.rewards.deterministic import dynamic_tracking_error

# %%
# Set the logger to use Comet
timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
api_key = "o77J6VCMDamustkfJuMXZ2jdV"
logger = CometLogger(
    api_key=api_key,
    project="RPA_6species_8rxns_MAK_REINFORCE",        
    workspace="maurice-filo", 
    name=f'RPA_6species_8rxns_MAK_REINFORCE_{timestamp}',
)
logger = logger.experiment

# %%
# Construct the template CRN
r1 = MassAction(reactant_labels=[], product_labels=['Z_1'], input_channels=['u_1'], params=[1.], params_controllability=[True])
r2 = MassAction(reactant_labels=['X_1'], product_labels=[], input_channels=['u_2'], params=[1.], params_controllability=[True])
r3 = MassAction(reactant_labels=['X_1'], product_labels=['X_1', 'X_2'], input_channels=[None], params=[1.], params_controllability=[True])
r4 = MassAction(reactant_labels=['X_2'], product_labels=[], input_channels=['u_3'], params=[1.], params_controllability=[True])
crn_template = IOCRN([r1, r2, r3, r4], output_labels=['X_2'])
crn_template.compile()
p = crn_template.num_inputs # Number of inputs of the IOCRNs
print("Template CRN:")
print(crn_template)

# Construct the library of possible reactions
species_labels = ['X_1', 'X_2', 'Z_1', 'Z_2', 'Z_3', 'Z_4']
library = construct_mass_action_library(species_labels=species_labels, order=2)
crn_template.set_library_context(library)
M = len(library.reactions) # Number of possible reactions
K = library.get_num_parameters() # Total number of parameters in all the reactions of the library
print("Library of possible reactions:")
print(library)
print("------------------------------------------------")

# %%
# Device
device = 'cuda' if torch.cuda.is_available() else 'cpu'
print(f'Using device: {device}')
print(f'Number of CPUs available: {os.cpu_count()}') 

# %%
# Flags and filenames
save_flag = True                                                # Save the agent checkpoint
load_flag = False                                               # Load the agent checkpoint 
train_flag = True                                               # Train the agent
save_sheet_flag = True                                          # Save the configuration to an Excel sheet

save_filename = timestamp + '.pth'                              # Filename for saving the agent checkpoint
load_filename = ''                                              # Filename for loading the agent checkpoint
file_name = "RPA_6species_8rxns_MAK_REINFORCE.xlsx"             # Filename for saving the Excel sheet

# %%
# Hyperparameters
max_added_reactions = 8                                 # Maximum number of reactions
N_CPUs = os.cpu_count()                                 # Number of CPUs          
N = 10*N_CPUs                                            # Number of samples (batch size)    
width = 1024                                            # Width of the neural networks  
depth = 5                                               # Depth of the neural networks 
deep_layer_size = 1024*10                               # Size of the deep layer encoding the CRNs
allow_input_influence = False                           # Allow input influence in the policy
learning_rate = 1e-4                                    # Learning rate for the optimizer 
hall_of_fame_size = 30                                  # Size of the hall of fame  
entropy_scheduler = {                                   # Entropy scheduler parameters 
    'entropy_weight': 1e-3 ,                            # Global entropy weight
    'entropy_weight_structure_head': 1.0,               # Entropy weight for the structure head
    'entropy_weight_continuous_head': 0.5,              # Entropy weight for the continuous parameters head
    'topk_entropy_weight': 1.0,                         # Entropy weight for the top-k actions
    'remainder_entropy_weight': 1.0,                    # Entropy weight for the remainder actions
    'entropy_update_coefficient': 1,                    # Entropy update coefficient (multiplicative)
    'entropy_schedule': 1000,                           # Entropy schedule (in epochs) 
    'minimum_entropy_weight': 0.0                       # Minimum entropy weight
}
structure_head_temperature = {
    'target_entropy_ratio_to_max': np.log(5)/np.log(M), 
    'initial_temperature': 1.0, 
    'rate': 0.0, 
    'current_temperature': 1.0}
risk_scheduler = {                                      # Risk scheduler parameters
    'risk': 0.9,                                        # Percentage of worst samples to disregard when computing the reward
    'risk_update': 0.0,                                 # Risk update coefficient (additive)
    'max_risk': 1.0,                                    # Maximum risk
    'risk_schedule': 1000                               # Risk schedule (in epochs)
}
epoch_num = 500                                         # Number of epochs for training
render_schedule = 10                                    # Render every # of epochs
render_mode = {                                         # Mode of the experiment
    'style': 'logger',                                  
    'task': 'transients', 
    'format': 'image',
    'topology': True,
    'bounds': [2.5]
}
render_n_best = 3                                       # Number of best CRNs to plot responses for
render_disregard_percentage = risk_scheduler['risk']    # Percentage of worst CRNs to disregard in the responses plotting

# Parameter distribution for the reactions added by the agent
continuous_distribution = {"type": 'lognormal_1D'}        

# Time horizon for the simulation
t_f = 100                                               # Final time for the simulation
N_t = 1000                                              # Number of time steps in the simulation
time_horizon = np.linspace(0, t_f, N_t, dtype=np.float32)

# Construct the IOCRN inputs
nums = [0.5, 1.0, 1.5]
u_list = [np.array(u) for u in product(nums, repeat=p)] # list of input combinations, each input is a numpy array of shape (p,)

# Construct the reference setpoints
r_list = [np.array([u[0]]) for u in u_list]

# Construct the IOCRN initial conditions
ic = IC(names=species_labels, values=[[0.01, 0.01, 0.01, 0.01, 0.01, 0.01]])

# Construct the weights for the performance metric
w = np.ones(N_t)
w[(len(w)//5)*4:] = w[(len(w)//5)*4:]*2
w[:(len(w)//5)] = w[:(len(w)//5)]*0.25
w = w[np.newaxis, :]

# Construct the compute reward routine
def compute_reward(state):
    x0_list = ic.get_ic(state)
    return dynamic_tracking_error(state, u_list, x0_list, time_horizon, r_list, w, norm=1, LARGE_NUMBER=1e4)

# %%
# Log the code and hyperparameters
# Log the code of the current file
current_file_path = os.path.abspath(__file__)
logger.log_code(file_name=os.path.basename(current_file_path))

# Log the hyperparameters
hyperparameters = {
    'species_labels': str(species_labels),
    'crn_template': str(crn_template),
    'library': str(library),
    'max_added_reactions': max_added_reactions,
    'N_CPUs': N_CPUs,
    'N': N,
    'width': width,
    'depth': depth,
    'deep_layer_size': deep_layer_size,
    'allow_input_influence': allow_input_influence,
    'learning_rate': learning_rate,
    'hall_of_fame_size': hall_of_fame_size,
    'entropy_scheduler': entropy_scheduler,
    'structure_head_temperature': structure_head_temperature,
    'risk_scheduler': risk_scheduler,
    'epoch_num': epoch_num,
    'render_schedule': render_schedule,
    'render_mode': render_mode,
    'render_n_best': render_n_best,
    'render_disregard_percentage': render_disregard_percentage,
    'continuous_distribution': continuous_distribution,
    'tf': t_f,
    'N_t': N_t,
    'u_list': [ u.tolist() for u in u_list ],
    'r_list': r_list,
    'ic': str(ic),
    'w': w.tolist()
}
logger.log_parameters(hyperparameters)

# %%
# Save the experiment configuration to an Excel sheet
if save_sheet_flag:
    sheet_name = "Data"
    headers = [
        "Timestamp", "URL",
        "Successful", "Epochs Completed", "Saved", "Comments",
        "Entropy Scheduler",
        "Learning Rate", "Epochs #",
        "(m, n, p, N)",
        "NN Depth", "NN Width", "Deep Layer Size", "CPUs #",
        "Risk Scheduler",
        "Render Schedule", "HoF Size",
        "Simulation Time", "Time Steps #",
        "Initial Conditions #", "Input Scenarios#",
        "Continuous Distribution", 
        "Structure Head Temperature"
    ]

    data_row = [
        timestamp, logger.url,
        None, None, None, None,
        str(entropy_scheduler),
        str(learning_rate), epoch_num,
        str((max_added_reactions, len(species_labels), p, N)),
        depth, width, deep_layer_size, N_CPUs,
        str(risk_scheduler),
        render_schedule, hall_of_fame_size,
        t_f, N_t, len(ic.values), len(u_list),
        str(continuous_distribution), 
        str(structure_head_temperature)
    ]

    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    # Write headers if sheet is empty
    if ws.max_row == 1 and ws.max_column == 1 and ws.cell(row=1, column=1).value is None:
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

    # Append experiment as next row
    next_row = ws.max_row + 1
    for col, value in enumerate(data_row, start=1):
        ws.cell(row=next_row, column=col, value=value)

    # Freeze header row and add filter
    ws.freeze_panes = "B1" 
    ws.auto_filter.ref = ws.dimensions

    # === Auto-fit column widths (except URL column) ===
    # URL column is column 2 (B), we leave its width unchanged.
    url_col_index = 2

    for col in range(1, ws.max_column + 1):
        if col == url_col_index:
            continue  # keep URL column width as-is

        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            value = cell.value
            if value is not None:
                # Convert to string to measure length
                length = len(str(value))
                if length > max_length:
                    max_length = length

        # Some padding so text isn't touching the cell border
        adjusted_width = max_length + 1 if max_length > 0 else 10
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(file_name)
    print(f"New experiment data saved in row {next_row} of '{file_name}'.")

# %%
# Construct parallel environments
crn_0 = crn_template.clone()
mult_env = ParallelEnvironments([Environment(crn_0, max_added_reactions, logger=logger, logger_schedule=1) for _ in range(N)], hall_of_fame_size=hall_of_fame_size, N_CPUs=N_CPUs, logger=logger)
# mult_env = SerialEnvironments([Environment(crn_0, max_added_reactions, logger=logger, logger_schedule=1) for _ in range(N)], hall_of_fame_size=hall_of_fame_size, logger=logger)

# %%
# Construct the policy
encoder_attributes = {"hidden_size": width, "num_layers": depth}
structure_head_attributes = {"hidden_size": width, "num_layers": depth}
rate_head_attributes = {"hidden_size": width, "num_layers": depth}
input_influence_head_attributes = {"hidden_size": width, "num_layers": depth}
masks = {"continuous": library.get_parameter_mask(mode="continuous"), "discrete": library.get_parameter_mask(mode="discrete"), "logit": library.get_logit_mask()}
entropy_weights_per_head = {'structure': entropy_scheduler['entropy_weight_structure_head'], 'continuous': entropy_scheduler['entropy_weight_continuous_head'], 'discrete': 0.0, 'input_influence': 0.0} 
policy = AddReactionByIndex(M, K, p, encoder_attributes, deep_layer_size, structure_head_attributes, rate_head_attributes, input_influence_head_attributes, allow_input_influence=allow_input_influence, masks=masks, device=device, continuous_distribution=continuous_distribution, entropy_weights_per_head=entropy_weights_per_head)

# Construct the agent
agent = REINFORCEAgent(policy, allow_input_influence=allow_input_influence, logger=logger, learning_rate=learning_rate, entropy_scheduler=entropy_scheduler, risk_scheduler=risk_scheduler, device=device)
if load_flag:
    agent.policy.load_state_dict(torch.load(load_filename+'.pth', map_location=device))

# %%
# Construct the interfaces
observer = ExplicitObserver(reaction_library=library, allow_input_observation=allow_input_influence)
tensorizer = ExplicitTensorizer(device=device)
actuator = LibraryActuator(reaction_library=library)
stepper = IOCRNStepper()

# %%
# Training Loop   
if train_flag:
    agent.policy.train()
    for i in tqdm(range(epoch_num)):
        mult_env.reset()
        for j in range(max_added_reactions):
            observations = mult_env.observe(observer, tensorizer)
            actions = agent.act(observations, actuator)
            out = mult_env.step(actions, stepper)
        rewards = mult_env.get_reward(compute_reward)
        agent.update(rewards, step_iteration=i)
        if i % render_schedule == 0:
            mult_env.render(rewards, n_best=render_n_best, disregarded_percentage=render_disregard_percentage, mode=render_mode)

# %%
# Test the model
agent.policy.eval()
mult_env.reset()
for j in range(max_added_reactions):
    observations = mult_env.observe(observer, tensorizer)
    actions = agent.act(observations, actuator)
    out = mult_env.step(actions, stepper)
rewards = mult_env.get_reward(compute_reward)

# Gather the CRNs from the environments
crns = mult_env.gather()

# Sort the CRNs by rewards
sorted_crns_rewards = sorted(zip(crns, rewards), key=lambda x: x[1])

# %%
# Plot the results
n_plot = 100
ax = None
for i in range(n_plot):
    x0_list = ic.get_ic(sorted_crns_rewards[i][0])
    time_horizon, x_list, y_list, last_task_info = sorted_crns_rewards[i][0].transient_response(u_list, x0_list, time_horizon)
    fig, ax = sorted_crns_rewards[i][0].plot_transient_response(axes=ax)
    print(f"IOCRN {i}, Reward: {sorted_crns_rewards[i][1]}")
    print(sorted_crns_rewards[i][0])

# %%
# Save the agent and the hall of fame CRNs
hall_of_fame_crns = [env.state for env in mult_env.hall_of_fame]
if save_flag:
    if not os.path.exists('models'):
        os.makedirs('models')
    if not os.path.exists('hof'):
        os.makedirs('hof')
    torch.save(agent.policy.state_dict(), 'models/' + save_filename)
    torch.save(hall_of_fame_crns, 'hof/hall_of_fame_' + save_filename)


